from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.db.models import Sum
from django.contrib.auth.models import User
from mkopoApp.models import Loan, LoanPayment
from loginApp.models import PaymentScreenshot
from mtajiApp.models import Mtaji
from michangoApp.models import Michango
from swadaqaApp.models import Swadaqa
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.db.models import Sum, Case, When
from django.db.models import Sum, Case, When, Value, DecimalField
from django.contrib.auth import get_user_model
from adminApp.models import PendingChanges, ActivityLog, VerifiedChanges, RejectedChanges
from django.utils import timezone
import openpyxl
from django.http import HttpResponse
from django.utils.timezone import now
from .models import MonthlyReport, YearlyReport
import os
from django.core.serializers.json import DjangoJSONEncoder
import json
from decimal import Decimal
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.db import transaction
from django.conf import settings
from .models import Notification
from django.urls import reverse
from .forms import AddMemberForm
from loginApp.models import Profile  # Import Profile model
from decimal import Decimal, InvalidOperation
def admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        # Authenticate the user with username and password
        user = authenticate(request, username=username, password=password)
        
        # Check if the user is valid and is a staff member
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('admin_App:admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials or not authorized.')
    
    return render(request, 'adminApp/admin_login.html', context={'is_login_page': True})


def admin_dashboard(request):
    context = get_dashboard_context()
    # Fetch unread notifications count for the logged-in admin
    unread_notifications_count = Notification.objects.filter(admin=request.user, is_read=False).count()
    
    # Add unread notifications count to the context
    context['unread_notifications_count'] = unread_notifications_count
    
    context['show_back_button'] = False
    return render(request, 'adminApp/dashboard.html', context)

def get_dashboard_context():
    total_members = User.objects.filter(is_staff=False, is_superuser=False).count()
    total_mtaji = Mtaji.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_michango = Michango.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_swadaqa = Swadaqa.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_loans = Loan.objects.filter(status='Approved').aggregate(Sum('amount'))['amount__sum'] or 0
    loan_requests = Loan.objects.filter(status='Pending').count()
    approved_loans = Loan.objects.filter(status='Approved').count()
    rejected_loans = Loan.objects.filter(status='Rejected').count()

    context = {
        'total_members': total_members,
        'total_mtaji': total_mtaji,
        'total_michango': total_michango,
        'total_swadaqa': total_swadaqa,
        'total_loans': total_loans,
        'loan_requests': loan_requests,
        'approved_loans': approved_loans,
        'rejected_loans': rejected_loans,
    }
    return context

def superuser_login_view(request):
    if request.user.is_authenticated:
        return redirect('admin_App:admin_dashboard')
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_superuser:
                    login(request, user)
                    return redirect('admin_App:admin_dashboard')
                else:
                    messages.error(request, 'You are not a Super User')
            else:
                messages.error(request, 'Invalid Username or Password')
    return render(request, 'adminApp/admin_login.html')

# add new member

@login_required(login_url='/account/login/')
def add_member(request):
    if request.method == 'POST':
        form = AddMemberForm(request.POST)
        if form.is_valid():
            user = form.save()  # Save the User instance first
            
            # Ensure the Profile is created after the user is saved
            Profile.objects.get_or_create(user=user)
            
            messages.success(request, 'New member added successfully.')
            return redirect(reverse('admin_App:members'))
        else:
            messages.error(request, 'There was an error with the form. Please check the inputs.')
    else:
        form = AddMemberForm()
    
    return render(request, 'adminApp/add_member.html', {'form': form})


#members
@login_required(login_url='/account/login/')
def members(request):
    members = User.objects.filter(is_staff=False, is_superuser=False)
    return render(request, 'adminApp/members.html', {'users': members, 'show_back_button': False})

@login_required(login_url='/account/login/')
def manage_user(request, user_id):
    # Fetch the user by ID
    user = get_object_or_404(User, id=user_id)
    
    # Create display name in "First_Last" format
    display_name = f"{user.first_name}_{user.last_name}".strip()
    
    # Fetch all contributions for Mtaji
    mtaji = Mtaji.objects.filter(user=user)
    total_mtaji = mtaji.aggregate(Sum('amount'))['amount__sum'] or 0  # Calculate total Mtaji
    
    # Fetch all contributions for Michango
    michango = Michango.objects.filter(user=user)
    total_mchango = michango.aggregate(Sum('amount'))['amount__sum'] or 0 
    
    # Fetch all contributions for Swadaqa
    swadaqa = Swadaqa.objects.filter(user=user)
    total_swadaqa = swadaqa.aggregate(Sum('amount'))['amount__sum'] or 0 

    # Start with the 'loans' data
    loans = Loan.objects.filter(user=user)
    mtaji = Mtaji.objects.filter(user=user)
    michango = Michango.objects.filter(user=user)
    swadaqa = Swadaqa.objects.filter(user=user)
    
    # Check if it works fine with just loans
    context = {
        'managed_user': user,
        'display_name': display_name,
        'loans': loans,
        'mtaji': mtaji,
        'total_mtaji': total_mtaji,
        'michango': michango,
        'total_mchango': total_mchango,
        'swadaqa': swadaqa,
        'total_swadaqa': total_swadaqa,
        'show_back_button': True,
    }
    return render(request, 'adminApp/manage_dashboard.html', context)

@login_required(login_url='/account/login/')
def view_payment_images(request, user_id):
    user = get_object_or_404(User, id=user_id)
    payment_images = PaymentScreenshot.objects.filter(user=user).order_by('-uploaded_at')
    context = {
        'managed_user': user,
        'payment_images': payment_images,
        'show_back_button': True,
    }
    return render(request, 'adminApp/view_payment_images.html', context)    


@login_required(login_url='/account/login/')
def notifications(request):
    # Fetch only unread notifications for the admin
    notifications = Notification.objects.filter(admin=request.user, is_read=False).order_by('-created_at')

    context = {
        'notifications': notifications,  # Only contains unread notifications
        'show_back_button': True,
    }
    return render(request, 'adminApp/notifications.html', context)


@login_required(login_url='/account/login/')
def mark_notification_as_read(request, notification_id):
    # Get the notification by ID and ensure it's for the current admin
    notification = get_object_or_404(Notification, id=notification_id, admin=request.user)

    # Mark the notification as read
    notification.is_read = True
    notification.save()

    # Redirect to the view payment image page
    return redirect(reverse('admin_App:view_payment_images', args=[notification.payment_screenshot.user.id]))


@login_required(login_url='/account/login/')
def view_member_details(request, user_id):
    member = get_object_or_404(User, id=user_id)
    profile = member.profile  # Assuming you have a Profile model linked to User

    context = {
        'member': member,
        'profile': profile,
        'show_back_button': True,
    }
    return render(request, 'adminApp/view_member_details.html', context)


@login_required(login_url='/account/login/')
def delete_member(request, user_id):
    member = get_object_or_404(User, id=user_id)
    member.delete()
    messages.success(request, 'Member deleted successfully.')
    return redirect(reverse('admin_App:members'))

# Mtaji
@login_required(login_url='/account/login/')
def manage_mtaji(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    # Create display name in "First_Last" format
    display_name = f"{user.first_name}_{user.last_name}".strip()

    # Generate years from 2018 to the current year
    years = [year for year in range(2018, timezone.now().year + 1)]
    current_year = timezone.now().year
    selected_year = int(request.GET.get('year', current_year))

    try:
        mtaji_record = Mtaji.objects.get(user=user, year=selected_year)
        current_amount = mtaji_record.amount
    except Mtaji.DoesNotExist:
        current_amount = 0

    if request.method == 'POST':
         # Check if the current user is admin1 and block updates if true
        if request.user.username == 'admin1':
            messages.error(request, "You are not authorized to update Mtaji data.")
            return redirect('admin_App:manage_mtaji', user_id=user.id)
        
        # Get the amount from POST and convert it to a decimal
        amount = request.POST.get(f'amount_{selected_year}')
        if amount:
            # Remove any formatting before converting to a decimal
            amount = amount.replace(',', '')
            try:
                amount = Decimal(amount)
                mtaji, created = Mtaji.objects.update_or_create(
                    user=user, 
                    year=selected_year,
                    defaults={'amount': amount}
                )
                # Pass the modified_by user (the admin) to the save method
                mtaji.save(modified_by=request.user)
            except (ValueError, InvalidOperation):
                messages.error(request, "Invalid amount format. Please enter a valid number.")
                return redirect('admin_App:manage_mtaji', user_id=user.id)

            # Redirect to the same page with the selected year after saving
            return redirect(f'{reverse("admin_App:manage_mtaji", kwargs={"user_id": user.id})}?year={selected_year}')

    context = {
        'managed_user': user,
        'display_name': display_name,
        'years': years,
        'current_year': selected_year,
        'current_amount': current_amount,
        'show_back_button': True,
    }
    return render(request, 'adminApp/manage_mtaji.html', context)


@login_required
def mtaji_data(request, user_id, year):
    user = get_object_or_404(User, id=user_id)
    try:
        mtaji = Mtaji.objects.get(user=user, year=year)
        amount = mtaji.amount
    except Mtaji.DoesNotExist:
        amount = 0

    data = {
        'year': year,
        'amount': amount,
        'total': amount
    }

    return JsonResponse(data)
   
@login_required(login_url='/account/login/')
def manage_mchango(request, user_id):
    user = get_object_or_404(User, id=user_id)
    current_year = timezone.now().year
    
    # Create display name in "First_Last" format
    display_name = f"{user.first_name}_{user.last_name}".strip()
    
    # Fetch Michango data for the selected user and year
    selected_year = int(request.GET.get('year', current_year))
    michango_entries = Michango.objects.filter(user=user, year=selected_year)

    # Define months in Kiswahili
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    
    months_with_data = {
        month: michango_entries.filter(month=index+1).first().amount if michango_entries.filter(month=index+1).exists() else 0
        for index, month in enumerate(months)
    }

    context = {
        'managed_user': user,
        'display_name': display_name,
        'current_year': selected_year,
        'years': range(2018, timezone.now().year + 1),
        'months_with_data': months_with_data.items(),
        'show_back_button': True,
    }
    
    return render(request, 'adminApp/manage_mchango.html', context)

@login_required(login_url='/account/login/')
def save_mchango(request, user_id):
    user = get_object_or_404(User, id=user_id)
    current_year = int(request.GET.get('year', timezone.now().year))

    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']

    modified_months = []  # Track modified months

    if request.method == 'POST':
        if request.user.username == 'admin1':
            messages.error(request, "You are not authorized to update Michango data.")
            return redirect('admin_App:manage_mchango', user_id=user.id)
        
        try:
            with transaction.atomic():
                for month in months:
                    amount = request.POST.get(f'amount_{month}')
                    if amount:
                        cleaned_amount = int(amount.replace(',', ''))
                        month_number = months.index(month) + 1
                        try:
                            michango = Michango.objects.get(user=user, year=current_year, month=month_number)
                            # Check if the amount has actually changed
                            if michango.amount != cleaned_amount:
                                michango.amount = cleaned_amount
                                michango.save(modified_by=request.user)
                                modified_months.append(month)  # Only log changed months
                        except Michango.DoesNotExist:
                            # If it doesn't exist, create new data and log the change
                            Michango.objects.create(
                                user=user,
                                year=current_year,
                                month=month_number,
                                amount=cleaned_amount,
                                modified_by=request.user
                            )
                            modified_months.append(month)  # Log the new month
        except IntegrityError:
            messages.error(request, "An error occurred while saving changes. Please try again.")
            return redirect(f'{reverse("admin_App:manage_mchango", kwargs={"user_id": user.id})}?year={current_year}')

        return redirect(f'{reverse("admin_App:manage_mchango", kwargs={"user_id": user.id})}?year={current_year}')

    return redirect(f'{reverse("admin_App:manage_mchango", kwargs={"user_id": user.id})}?year={current_year}')


@login_required(login_url='/account/login/')
def manage_swadaqa(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    # Create display name in "First_Last" format
    display_name = f"{user.first_name}_{user.last_name}".strip()

    # Generate years from 2018 to the current year
    years = [year for year in range(2018, timezone.now().year + 1)]
    current_year = timezone.now().year
    selected_year = int(request.GET.get('year', current_year))

    try:
        swadaqa_record = Swadaqa.objects.get(user=user, year=selected_year)
        current_amount = swadaqa_record.amount
    except Swadaqa.DoesNotExist:
        current_amount = 0

    if request.method == 'POST':
        if request.user.username == 'admin1':
            messages.error(request, "You are not authorized to update Swadaqa data.")
            return redirect(f'{reverse("admin_App:manage_swadaqa", kwargs={"user_id": user.id})}?year={selected_year}')
        
        # Get the amount from POST and convert it to a decimal
        amount = request.POST.get(f'amount_{selected_year}')
        if amount:
            amount = amount.replace(',', '')  # Remove any commas before conversion
            try:
                amount = Decimal(amount)
                swadaqa, created = Swadaqa.objects.update_or_create(
                    user=user,
                    year=selected_year,
                    defaults={'amount': amount}
                )
                swadaqa.save(modified_by=request.user)
            except (ValueError, InvalidOperation):
                messages.error(request, "Invalid amount format. Please enter a valid number.")
                return redirect(f'{reverse("admin_App:manage_swadaqa", kwargs={"user_id": user.id})}?year={selected_year}')
            
            return redirect(f'{reverse("admin_App:manage_swadaqa", kwargs={"user_id": user.id})}?year={selected_year}')

    context = {
        'managed_user': user,
        'display_name': display_name,
        'years': years,
        'current_year': selected_year,
        'current_amount': current_amount,
        'show_back_button': True,
    }
    return render(request, 'adminApp/manage_swadaqa.html', context)



@login_required
def swadaqa_data(request, user_id, year):
    user = get_object_or_404(User, id=user_id)
    try:
        swadaqa = Swadaqa.objects.get(user=user, year=year)
        amount = swadaqa.amount
    except Swadaqa.DoesNotExist:
        amount = 0

    data = {
        'year': year,
        'amount': amount,
        'total': amount
    }
    
    return JsonResponse(data)

@login_required(login_url='/account/login/')
def manage_mkopo(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Step 1: Check loan status
    current_loan = Loan.objects.filter(user=user, status='Approved').first()  # Get the first approved loan, if any
    loan_status = 'Active' if current_loan else 'Inactive'

    # Step 2: Loan Limit Message
    loan_limit = 10000000  # Example loan limit

    # Step 3: Check if there's a pending loan request
    loan_request = Loan.objects.filter(user=user, status='Pending').first()  # Get the first pending loan, if any
    has_request = loan_request is not None

    # Step 4: Loan History
    loan_history = Loan.objects.filter(user=user).exclude(status='Pending')

# Step 5: Loan Payments and Total Payments
    if current_loan:
        loan_payments = LoanPayment.objects.filter(loan=current_loan).order_by('payment_date')
        loan_payments_total = loan_payments.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.0')
        remaining_loan_balance = current_loan.amount - loan_payments_total
    else:
        loan_payments = []
        loan_payments_total = Decimal('0.0')
        remaining_loan_balance = Decimal('0.0')

    context = {
        'managed_user': user,
        'loan_status': loan_status,
        'loan_limit': loan_limit,
        'has_request': has_request,
        'loan_request': loan_request,
        'loan_history': loan_history,
        'current_loan': current_loan,
        'loan_payments': loan_payments,
        'loan_payments_total': loan_payments_total,
        'remaining_loan_balance': remaining_loan_balance,
        'show_back_button': True,
    }

    return render(request, 'adminApp/manage_mkopo.html', context)


@login_required(login_url='/account/login/')
def process_loan_request(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        if request.user.username == 'admin1':
            messages.error(request, "You are not authorized to process loan requests.")
            return redirect('admin_App:manage_mkopo', user_id=user.id)
        
        decision = request.POST.get('decision')
        
        # Retrieve all pending loans for the user
        loan_requests = Loan.objects.filter(user=user, status='Pending')

        if loan_requests.exists():
            display_name = f"{user.first_name}_{user.last_name}".strip()
            # Process each loan request (if more than one, process each or decide on other logic)
            for loan_request in loan_requests:
                # Change loan status to 'Pending Verification'
                loan_request.status = 'Pending Verification'
                loan_request.save()

                # Prepare the data to be logged
                data = {
                    'user': loan_request.user.display_name,
                    'loan_id': loan_request.id,
                    'original_status': 'Pending',
                    'new_status': 'Approved' if decision == 'accept' else 'Rejected',
                }

                # Create a PendingChanges entry for verification
                if not PendingChanges.objects.filter(
                    admin=request.user,
                    table_name='Loan',
                    action='Update',
                    data=json.dumps(data, cls=DjangoJSONEncoder)
                ).exists():
                    PendingChanges.objects.create(
                        admin=request.user,
                        table_name='Loan',
                        action='Update',
                        data=json.dumps(data, cls=DjangoJSONEncoder)
                    )

    return redirect('admin_App:manage_mkopo', user_id=user.id)

@login_required(login_url='/account/login/')
def loan_payments_view(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Get the most recent approved loan for the user
    current_loan = Loan.objects.filter(user=user, status='Approved').order_by('-date').first()

    if not current_loan:
        messages.error(request, 'No approved loans found for this user.')
        return redirect('admin_App:manage_mkopo', user_id=user_id)
    
    # Get the total amount of payments already made for this loan
    total_payments = LoanPayment.objects.filter(loan=current_loan).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.0')

    if request.method == 'POST':
        if request.user.username == 'admin1':
            messages.error(request, "You are not authorized to add loan payments.")
            return redirect('admin_App:manage_mkopo', user_id=user_id)
        
        payment_date = request.POST.get('payment_date')
        amount = request.POST.get('amount')

        try:
            amount = Decimal(amount)
        except (ValueError, TypeError):
            messages.error(request, 'Invalid amount entered.')
            return redirect('admin_App:manage_mkopo', user_id=user_id)

        try:
            payment_date = datetime.strptime(payment_date, '%d/%m/%Y')
        except ValueError:
            messages.error(request, 'Invalid date format. Please use DD/MM/YYYY.')
            return redirect('admin_App:manage_mkopo', user_id=user_id)
        
        # Calculate the total payments after adding this new payment
        new_total_payments = total_payments + amount

        # Check if the new total payments exceed the loan amount
        if new_total_payments > current_loan.amount:
            messages.error(request, 'Payment exceeds the remaining loan balance.')
            return redirect('admin_App:manage_mkopo', user_id=user_id)

        # Create a new payment
        payment = LoanPayment.objects.create(
            loan=current_loan,
            payment_date=payment_date,
            amount=amount,
            modified_by=request.user
        )

        # Log the action (including the amount)
        data = {
            'user': user.username,
            'loan_id': current_loan.id,
            'action': 'Update',
            'amount': str(payment.amount),  # Add the amount to the log
        }

        PendingChanges.objects.create(
            admin=request.user,
            table_name='Loan',
            action='Update',
            data=json.dumps(data, cls=DjangoJSONEncoder)
        )

        messages.success(request, 'Payment recorded successfully.')
        return redirect('admin_App:manage_mkopo', user_id=user_id)

    return redirect('admin_App:manage_mkopo', user_id=user_id)


@login_required(login_url='/account/login/')
def mitaji_view(request):
    current_year = datetime.now().year
    years = list(range(2018, current_year + 1))
    selected_year = int(request.GET.get('year', current_year))

    total_mtaji = Mtaji.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_mtaji_year = Mtaji.objects.filter(year=selected_year).aggregate(Sum('amount'))['amount__sum'] or 0
    members = User.objects.filter(is_staff=False, is_superuser=False)
    
    members_mtaji = []
    for member in members:
        # Calculate total amount for all years        
        mtaji = Mtaji.objects.filter(user=member, year=selected_year).first()
        amount = mtaji.amount if mtaji else 0
        display_name = f"{member.first_name}_{member.last_name}".strip()
        members_mtaji.append({
            'user_id': member.id,
            'display_name': display_name,
            'amount': amount,
        })

    context = {
        'total_mtaji': total_mtaji,
        'total_mtaji_year': total_mtaji_year,
        'years': years,
        'selected_year': selected_year,
        'members_mtaji': members_mtaji
    }

    return render(request, 'adminApp/mitaji.html', context)


#Michango
@login_required
def michango_view(request):
    current_year = datetime.now().year
    years = list(range(2018, current_year + 1))
    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    
    # Get selected year from the request (default to current year)
    selected_year = int(request.GET.get('year', current_year))

    # Fetch total michango for all years
    total_michango = Michango.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    
     # Fetch total michango for the selected year
    total_year_michango = Michango.objects.filter(year=selected_year).aggregate(Sum('amount'))['amount__sum'] or 0
    
     # Fetch michango data for each month in the selected year
    total_michango_by_month = [Michango.objects.filter(year=selected_year, month=i+1).aggregate(Sum('amount'))['amount__sum'] or 0 for i in range(12)]

    members = User.objects.filter(is_staff=False, is_superuser=False)
    members_michango = []
    for member in members:
        member_michango_by_month = [Michango.objects.filter(user=member, year=selected_year, month=i+1).aggregate(Sum('amount'))['amount__sum'] or 0 for i in range(12)]
        display_name = f"{member.first_name}_{member.last_name}".strip()
        members_michango.append({
            'user_id': member.id,
            'display_name': display_name,
            'michango_by_month': member_michango_by_month
        })

    context = {
        'total_michango': total_michango,
        'total_year_michango': total_year_michango,
        'total_michango_by_month': total_michango_by_month,
        'years': years,
        'selected_year': selected_year,
        'months': months,
        'members_michango': members_michango
    }

    return render(request, 'adminApp/michango.html', context)
@login_required(login_url='/account/login')
def swadaqa_view(request):
    current_year = datetime.now().year
    years = list(range(2018, current_year + 1))
    selected_year = int(request.GET.get('year', current_year))

    total_swadaqa = Swadaqa.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_swadaqa_year = Swadaqa.objects.filter(year=selected_year).aggregate(Sum('amount'))['amount__sum'] or 0
    members = User.objects.filter(is_staff=False, is_superuser=False)

    members_swadaqa = []
    for member in members:
        swadaqa = Swadaqa.objects.filter(user=member, year=selected_year).first()
        amount = swadaqa.amount if swadaqa else 0
        display_name = f"{member.first_name}_{member.last_name}".strip()
        members_swadaqa.append({
            'user_id': member.id,
            'display_name': display_name,
            'amount': amount
        })

    context = {
        'total_swadaqa': total_swadaqa,
        'total_swadaqa_year': total_swadaqa_year,
        'years': years,
        'selected_year': selected_year,
        'members_swadaqa': members_swadaqa
    }

    return render(request, 'adminApp/swadaqas.html', context)

#loans
@login_required(login_url='/account/login/')
def members_loans(request):
    current_year = datetime.now().year
    selected_year = int(request.GET.get('year', current_year))
    
    # Get total loans for the selected year
    total_year_loans = Loan.objects.filter(status='Approved', date__year=selected_year).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Get total loans for all years
    total_loans_all_years = Loan.objects.filter(status='Approved').aggregate(Sum('amount'))['amount__sum'] or 0

    members = User.objects.filter(is_staff=False, is_superuser=False).annotate(
        total_loan=Sum(
            Case(
                When(loan__status='Approved', loan__date__year=selected_year, then='loan__amount'),
                default=Value(0),
                output_field=DecimalField()
            )
        )
    )

    members_loans_info = []
    for member in members:
        loan = Loan.objects.filter(user=member, status='Approved', date__year=selected_year).first()
        loan_status = 'No Loan' if loan is None else 'Approved'
        loan_amount = 0 if loan is None else loan.amount
        display_name = f"{member.first_name}_{member.last_name}".strip()
        members_loans_info.append({
            'user_id': member.id,  # Ensure user_id is included
            'display_name': display_name,
            'phone_number': member.profile.phone_number,
            'loan_status': loan_status,
            'loan_amount': loan_amount
        })

    context = {
        'members_loans_info': members_loans_info,
        'total_loans': total_loans_all_years,  # Total of all years
        'total_year_loans': total_year_loans,  # Total for the selected year
        'years': list(range(2018, current_year + 1)),
        'selected_year': selected_year,
    }

    return render(request, 'adminApp/members_loans.html', context)

@login_required(login_url='/account/login/')
def loan_requests(request):
    loan_requests = Loan.objects.filter(status='Pending')
    return render(request, 'adminApp/loan_requests.html', {'loans': loan_requests, 'show_back_button': False})

@login_required(login_url='/account/login/')
def approved_loans(request):
    approved_loans = Loan.objects.filter(status='Approved')
    return render(request, 'adminApp/approved_loans.html', {'loans': approved_loans, 'show_back_button': False})

@login_required(login_url='/account/login/')
def rejected_loans(request):
    rejected_loans = Loan.objects.filter(status='Rejected')
    return render(request, 'adminApp/rejected_loans.html', {'loans': rejected_loans, 'show_back_button': False})

@login_required(login_url='/account/login/')
def approve_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)

    # Change loan status to 'Pending Verification'
    loan.status = 'Pending Verification'
    loan.save()

    # Prepare the data to be logged
    data = {
        'user': loan.user.username,
        'loan_id': loan.id,
        'amount': str(loan.amount),  # Add the loan amount here
        'original_status': 'Pending',
        'new_status': 'Approved',
    }

    # Create a PendingChanges entry for verification
    if not PendingChanges.objects.filter(
        admin=request.user,
        table_name='Loan',
        action='Update',
        data=json.dumps(data, cls=DjangoJSONEncoder)
    ).exists():
        PendingChanges.objects.create(
            admin=request.user,
            table_name='Loan',
            action='Update',
            data=json.dumps(data, cls=DjangoJSONEncoder)
        )

    messages.success(request, 'Loan approval submitted for verification.')
    return redirect('admin_App:loan_requests')


@login_required(login_url='/account/login/')
def reject_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)

    # Change loan status to 'Pending Verification'
    loan.status = 'Pending Verification'
    loan.save()

    # Prepare the data to be logged
    data = {
        'user': loan.user.username,
        'loan_id': loan.id,
        'amount': str(loan.amount),  # Add the loan amount here
        'original_status': 'Pending',
        'new_status': 'Rejected',
    }

    # Create a PendingChanges entry for verification
    if not PendingChanges.objects.filter(
        admin=request.user,
        table_name='Loan',
        action='Update',
        data=json.dumps(data, cls=DjangoJSONEncoder)
    ).exists():
        PendingChanges.objects.create(
            admin=request.user,
            table_name='Loan',
            action='Update',
            data=json.dumps(data, cls=DjangoJSONEncoder)
        )

    messages.success(request, 'Loan rejection submitted for verification.')
    return redirect('admin_App:loan_requests')


import json
# verification for admin1
@login_required(login_url='/account/login/')
def verification(request):
    if request.user.username == 'admin1':
        # Fetch all pending changes that have not been approved yet, ordered by the latest created date.
        pending_changes = PendingChanges.objects.filter(is_approved=False).order_by('-created_at')

        # Ensure that the data field is properly parsed if it's stored as JSON.
        for change in pending_changes:
            if isinstance(change.data, str):
                change.data = json.loads(change.data)
           
            # Create display name in "First_Last" format if user data is present
            if 'user' in change.data and change.data['user']:
                user_identifier = change.data['user']     
           # Check if `user_identifier` is a digit (an ID) or a username
                if str(user_identifier).isdigit():
                    user = User.objects.filter(id=user_identifier).first()
                else:
                    user = User.objects.filter(username=user_identifier).first()

                if user:
                    change.data['display_name'] = f"{user.first_name}_{user.last_name}".strip()
                else:
                    change.data['display_name'] = "Unknown User"

        # Render the verification page with the pending changes.
        return render(request, 'adminApp/verification.html', {'pending_changes': pending_changes})
    else:
        # If the user is not 'admin1', deny access and redirect to the admin dashboard.
        messages.error(request, 'Access denied.')
        return redirect('admin_App:admin_dashboard')
    
# Add this MONTHS_MAP here:
MONTHS_MAP = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4, 
    'May': 5, 'June': 6, 'July': 7, 'August': 8, 
    'September': 9, 'October': 10, 'November': 11, 'December': 12
}

@login_required(login_url='/account/login/')
def approve_change(request, change_id):
    if request.user.username == 'admin1':
        pending_change = get_object_or_404(PendingChanges, action_no=change_id)
        pending_change.is_approved = True
        pending_change.approved_by = request.user
        pending_change.save()

        # Ensure the data is parsed as a dictionary
        if isinstance(pending_change.data, str):
            pending_change.data = json.loads(pending_change.data)

        # Handle different categories: Loan, Mtaji, Michango, Swadaqa
        if pending_change.table_name == 'Loan':
            loan = get_object_or_404(Loan, id=pending_change.data.get('loan_id'))
            loan.status = pending_change.data.get('new_status', 'Approved')
            loan.save()
        elif pending_change.table_name == 'Mtaji':
            user = get_object_or_404(User, username=pending_change.data.get('user'))
            Mtaji.objects.update_or_create(
                user=user,
                year=pending_change.data.get('year'),
                defaults={'amount': pending_change.data.get('amount')}
            )
        elif pending_change.table_name == 'Michango':
            user = get_object_or_404(User, username=pending_change.data.get('user'))
            Michango.objects.update_or_create(
                user=user,
                year=pending_change.data.get('year'),
                month=pending_change.data.get('month'),
                defaults={'amount': pending_change.data.get('amount')}
            )
        elif pending_change.table_name == 'Swadaqa':
            user = get_object_or_404(User, username=pending_change.data.get('user'))
            Swadaqa.objects.update_or_create(
                user=user,
                year=pending_change.data.get('year'),
                defaults={'amount': pending_change.data.get('amount')}
            )

        # Log the action in ActivityLog
        ActivityLog.objects.create(
            admin=request.user,
            action=f"Approved {pending_change.action}",
            affected_user=pending_change.data.get('user'),
            amount=pending_change.data.get('amount'),
            details=f"Approved changes to {pending_change.table_name}"
        )

        # Store the approved change in VerifiedChanges
        VerifiedChanges.objects.create(pending_change=pending_change)

        messages.success(request, 'Change approved successfully.')

    return redirect('admin_App:verification')
def revert_mtaji_change(pending_change):
    # Extract the data from the pending change
    data = pending_change.data
    user = User.objects.get(username=data['user'])
    year = data.get('year')
    
    # Revert the change
    if pending_change.action == 'Create':
        # If it was a creation, delete the record
        Mtaji.objects.filter(user=user, year=year).delete()
    elif pending_change.action == 'Update':
        # If it was an update, revert to the previous value if available
        original_amount = data.get('original_amount', 0)
        Mtaji.objects.filter(user=user, year=year).update(amount=original_amount)
        
def revert_michango_change(pending_change):
    # Extract the data from the pending change
    data = pending_change.data
    user = User.objects.get(username=data['user'])
    year = data.get('year')
    month = data.get('month')
    
    # Revert the change
    if pending_change.action == 'Create':
        # If it was a creation, delete the record
        Michango.objects.filter(user=user, year=year, month=month).delete()
    elif pending_change.action == 'Update':
        # If it was an update, revert to the previous value if available
        original_amount = data.get('original_amount', 0)
        Michango.objects.filter(user=user, year=year, month=month).update(amount=original_amount)
        
        
def revert_swadaqa_change(pending_change):
    # Extract the data from the pending change
    data = pending_change.data
    user = User.objects.get(username=data['user'])
    year = data.get('year')
    
    # Revert the change
    if pending_change.action == 'Create':
        # If it was a creation, delete the record
        Swadaqa.objects.filter(user=user, year=year).delete()
    elif pending_change.action == 'Update':
        # If it was an update, revert to the previous value if available
        original_amount = data.get('original_amount', 0)
        Swadaqa.objects.filter(user=user, year=year).update(amount=original_amount)


def revert_loan_change(pending_change):
    # Extract the data from the pending change
    data = pending_change.data
    user = User.objects.get(username=data['user'])
    loan_id = data.get('loan_id')

    # Revert the change
    if pending_change.action == 'Create':
        # If it was a creation, delete the loan record
        Loan.objects.filter(id=loan_id, user=user).delete()
    elif pending_change.action == 'Update':
        # If it was an update, revert to the previous status if available
        original_status = data.get('original_status', 'Pending')
        Loan.objects.filter(id=loan_id, user=user).update(status=original_status)


@login_required(login_url='/account/login/')
def reject_change(request, change_id):
    if request.user.username == 'admin1':
        try:
            with transaction.atomic():
                pending_change = get_object_or_404(PendingChanges, action_no=change_id)

                # Ensure the data is parsed as a dictionary
                if isinstance(pending_change.data, str):
                    pending_change.data = json.loads(pending_change.data)

                # Handle different categories: Loan, Mtaji, Michango, Swadaqa
                if pending_change.table_name == 'Loan':
                    loan = get_object_or_404(Loan, id=pending_change.data.get('loan_id'))
                    loan.status = pending_change.data.get('original_status', 'Pending')
                    loan.save()
                elif pending_change.table_name == 'Mtaji':
                    revert_mtaji_change(pending_change)
                elif pending_change.table_name == 'Michango':
                    revert_michango_change(pending_change)
                elif pending_change.table_name == 'Swadaqa':
                    revert_swadaqa_change(pending_change)

                # Store the rejected change in RejectedChanges
                rejected_change = RejectedChanges.objects.create(pending_change=pending_change)

                # Log the rejection action in ActivityLog
                ActivityLog.objects.create(
                    admin=request.user,
                    action=f"Rejected {pending_change.action}",
                    affected_user=pending_change.data.get('user'),
                    amount=pending_change.data.get('amount'),
                    details=f"Rejected changes to {pending_change.table_name}"
                )

                # Now delete the pending change after rejection has been logged
                pending_change.delete()

            messages.success(request, 'Change rejected successfully.')

        except Exception as e:
            print(f"An error occurred: {e}")
            messages.error(request, f"An error occurred while rejecting the change: {e}")

    return redirect('admin_App:verification')


@login_required(login_url='/account/login/')
def verified_actions(request):
    if request.user.is_staff:
        # Fetch all verified changes, ordered by the latest verified date.
        verified_changes = VerifiedChanges.objects.select_related('pending_change').order_by('-verified_at')

        # Parse the data in each verified change
        for change in verified_changes:
            if isinstance(change.pending_change.data, str):
                change.pending_change.data = json.loads(change.pending_change.data)
            
             # Create display name in "First_Last" format if user data is present
            if 'user' in change.pending_change.data and change.pending_change.data['user']:
                user_identifier = change.pending_change.data['user']

                # Check if `user_identifier` is a digit (an ID) or a username
                if str(user_identifier).isdigit():
                    user = User.objects.filter(id=user_identifier).first()
                else:
                    user = User.objects.filter(username=user_identifier).first()

                if user:
                    change.pending_change.data['display_name'] = f"{user.first_name}_{user.last_name}".strip()
                else:
                    change.pending_change.data['display_name'] = "Unknown User"

        # Render the verified actions page with the verified changes.
        return render(request, 'adminApp/verified_actions.html', {'verified_actions': verified_changes})
    else:
        # If the user is not 'admin1', deny access and redirect to the admin dashboard.
        messages.error(request, 'Access denied.')
        return redirect('admin_App:admin_dashboard')
    

@login_required(login_url='/account/login/')
def rejected_actions(request):
    if request.user.is_staff:
        # Retrieve all rejected actions where pending_change is not None, ordered by the latest rejected date.
        rejected_actions = RejectedChanges.objects.filter(pending_change__isnull=False).select_related('pending_change').order_by('-rejected_at')

        # Parse the data in each rejected change
        for change in rejected_actions:
            if isinstance(change.pending_change.data, str):
                change.pending_change.data = json.loads(change.pending_change.data)
            
            # Create display name in "First_Last" format if user data is present
            if 'user' in change.pending_change.data and change.pending_change.data['user']:
                user_identifier = change.pending_change.data['user']

                # Check if `user_identifier` is a digit (an ID) or a username
                if str(user_identifier).isdigit():
                    user = User.objects.filter(id=user_identifier).first()
                else:
                    user = User.objects.filter(username=user_identifier).first()

                if user:
                    change.pending_change.data['display_name'] = f"{user.first_name}_{user.last_name}".strip()
                else:
                    change.pending_change.data['display_name'] = "Unknown User"

        # Render the rejected actions page with the rejected changes.
        return render(request, 'adminApp/rejected_actions.html', {'rejected_actions': rejected_actions})
    else:
        # If the user is not 'admin1', deny access and redirect to the admin dashboard.
        messages.error(request, 'Access denied.')
        return redirect('admin_App:admin_dashboard')
    
@login_required(login_url='/account/login/')
def report_dashboard(request):
    # Retrieve and order reports
    monthly_reports = MonthlyReport.objects.all().order_by('-year', '-month')
    yearly_reports = YearlyReport.objects.all().order_by('-year')
    
    context = {
        'monthly_reports': monthly_reports,
        'yearly_reports': yearly_reports,
    }
    return render(request, 'adminApp/report_dashboard.html', context)

@login_required(login_url='/account/login/')
def view_monthly_report(request, id):
    report = get_object_or_404(MonthlyReport, id=id)
    filepath = report.file_path

    # Load the Excel workbook and active sheet
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Extract data from the sheet
    report_data = list(ws.iter_rows(values_only=True))

    context = {
        'report_data': report_data[1:],  # Skip headers
        'headers': report_data[0],       # The first row is the header
        'report': report,
    }

    return render(request, 'adminApp/view_monthly_report.html', context)

@login_required(login_url='/account/login/')
def view_yearly_report(request, id):
    report = get_object_or_404(YearlyReport, id=id)
    filepath = report.file_path

    # Load the Excel workbook and active sheet
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Extract data from the sheet
    report_data = list(ws.iter_rows(values_only=True))

    context = {
        'report_data': report_data[1:],  # Skip headers
        'headers': report_data[0],       # The first row is the header
        'report': report,
    }

    return render(request, 'adminApp/view_yearly_report.html', context)

@login_required(login_url='/account/login/')
def download_monthly_report(request, id):
    report = get_object_or_404(MonthlyReport, id=id)
    filepath = report.file_path
    with open(filepath, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(filepath)}'
        return response

@login_required(login_url='/account/login/')
def download_yearly_report(request, id):
    report = get_object_or_404(YearlyReport, id=id)
    filepath = report.file_path
    with open(filepath, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(filepath)}'
        return response